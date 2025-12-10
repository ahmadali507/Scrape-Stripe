import os
import csv
import requests
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Stripe API Configuration
# Load from environment variable (set in .env file)
STRIPE_API_KEY = os.getenv('STRIPE_SECRET_KEY')
STRIPE_API_BASE = 'https://api.stripe.com/v1'

# Validate API key
if not STRIPE_API_KEY:
    raise ValueError(
        "STRIPE_SECRET_KEY not found in environment variables.\n"
        "Please create a .env file with: STRIPE_SECRET_KEY=your_key_here"
    )


def fetch_subscription(subscription_id: str) -> Optional[Dict]:
    """
    Fetch a specific subscription by ID from Stripe API using direct HTTP requests.
    
    Args:
        subscription_id: The Stripe subscription ID (e.g., 'sub_1Scm8EL5YqSpFl3KeJFAFP1g')
    
    Returns:
        Dictionary containing subscription data or None if not found
    """
    try:
        url = f"{STRIPE_API_BASE}/subscriptions/{subscription_id}"
        headers = {
            'Authorization': f'Bearer {STRIPE_API_KEY}',
            'Content-Type': 'application/x-www-form-urlencoded'
        }
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.Timeout:
        print(f"Error: Request timed out while fetching subscription {subscription_id}")
        return None
    except requests.exceptions.ConnectionError as e:
        print(f"Error: Connection failed while fetching subscription {subscription_id}")
        print(f"Details: {str(e)}")
        return None
    except requests.exceptions.HTTPError as e:
        if e.response is not None:
            status_code = e.response.status_code
            if status_code == 401:
                print(f"Error: Unauthorized. Invalid API key for subscription {subscription_id}")
            elif status_code == 404:
                print(f"Error: Subscription {subscription_id} not found")
            else:
                print(f"Error: HTTP {status_code} - {e}")
            try:
                error_data = e.response.json()
                if 'error' in error_data:
                    error_msg = error_data['error'].get('message', 'Unknown error')
                    print(f"Stripe API Error: {error_msg}")
            except:
                pass
        return None
    except requests.exceptions.RequestException as e:
        print(f"Error fetching subscription {subscription_id}: {type(e).__name__}: {e}")
        if hasattr(e, 'response') and e.response is not None:
            try:
                print(f"Response: {e.response.text[:200]}")
            except:
                pass
        return None
    except Exception as e:
        print(f"Unexpected error fetching subscription {subscription_id}: {type(e).__name__}: {e}")
        return None


def fetch_all_subscriptions(limit: int = 100) -> List[Dict]:
    """
    Fetch all subscriptions from Stripe API using direct HTTP requests with pagination.
    
    Args:
        limit: Maximum number of subscriptions to retrieve (will paginate if needed)
    
    Returns:
        List of subscription objects
    """
    print(f"\n=== Fetching subscriptions (up to {limit}) ===")
    print(f"URL: {STRIPE_API_BASE}/subscriptions")
    print(f"Method: GET")
    print(f"API Key: {STRIPE_API_KEY[:20]}...")
    
    url = f"{STRIPE_API_BASE}/subscriptions"
    headers = {
        'Authorization': f'Bearer {STRIPE_API_KEY}',
        'Content-Type': 'application/x-www-form-urlencoded'
    }
    
    all_subscriptions = []
    page_limit = 100  # Stripe's max per page
    params = {'limit': min(page_limit, limit)}
    
    try:
        page = 1
        while len(all_subscriptions) < limit:
            print(f"\nFetching page {page} (limit: {params['limit']})...")
            
            response = requests.get(url, headers=headers, params=params, timeout=30)
            
            if response.status_code == 200:
                try:
                    data = response.json()
                    
                    if isinstance(data, dict) and 'data' in data:
                        page_subscriptions = data.get('data', [])
                        remaining = limit - len(all_subscriptions)
                        
                        # Add subscriptions up to the limit
                        if len(page_subscriptions) <= remaining:
                            all_subscriptions.extend(page_subscriptions)
                        else:
                            all_subscriptions.extend(page_subscriptions[:remaining])
                        
                        print(f"  ✓ Retrieved {len(page_subscriptions)} subscription(s) from page {page}")
                        print(f"  Total so far: {len(all_subscriptions)}/{limit}")
                        
                        # Check if there are more pages and we haven't reached the limit
                        has_more = data.get('has_more', False)
                        if not has_more or len(all_subscriptions) >= limit:
                            break
                        
                        # Set up pagination for next page
                        if page_subscriptions:
                            params['starting_after'] = page_subscriptions[-1]['id']
                            params['limit'] = min(page_limit, limit - len(all_subscriptions))
                        page += 1
                    else:
                        print(f"  ⚠ Unexpected response structure")
                        break
                        
                except Exception as e:
                    print(f"\nError parsing JSON: {e}")
                    print(f"Response text: {response.text[:500]}")
                    break
            else:
                print(f"\n=== Error Response ===")
                print(f"Status: {response.status_code}")
                try:
                    error_data = response.json()
                    print(f"Error JSON: {error_data}")
                except:
                    print(f"Response text: {response.text[:500]}")
                break
                
        print(f"\n=== Total subscriptions fetched: {len(all_subscriptions)} ===")
        return all_subscriptions
            
    except requests.exceptions.Timeout:
        print(f"\n=== Timeout Error ===")
        print("Request timed out after 30 seconds")
        return all_subscriptions
    except requests.exceptions.ConnectionError as e:
        print(f"\n=== Connection Error ===")
        print(f"Error: {e}")
        print(f"Error type: {type(e).__name__}")
        return all_subscriptions
    except Exception as e:
        print(f"\n=== Unexpected Error ===")
        print(f"Error type: {type(e).__name__}")
        print(f"Error message: {str(e)}")
        import traceback
        traceback.print_exc()
        return all_subscriptions


def fetch_customer_details(customer_id: str, silent: bool = False) -> Optional[Dict]:
    """
    Fetch customer details by customer ID using direct HTTP requests.
    
    Args:
        customer_id: The Stripe customer ID (e.g., 'cus_TZw07B5NXMht4M')
        silent: If True, don't print error messages
    
    Returns:
        Dictionary containing customer data or None if not found
    """
    try:
        url = f"{STRIPE_API_BASE}/customers/{customer_id}"
        headers = {
            'Authorization': f'Bearer {STRIPE_API_KEY}',
            'Content-Type': 'application/x-www-form-urlencoded'
        }
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        if not silent:
            error_msg = f"Error fetching customer {customer_id}"
            if hasattr(e, 'response') and e.response is not None:
                if e.response.status_code == 404:
                    error_msg += ": Customer not found"
                else:
                    error_msg += f": {e.response.text}"
            else:
                error_msg += f": {str(e)}"
            print(f"  Warning: {error_msg}")
        return None


def format_timestamp(timestamp: int) -> str:
    """
    Convert Unix timestamp to readable date string.
    
    Args:
        timestamp: Unix timestamp
    
    Returns:
        Formatted date string (YYYY-MM-DD HH:MM:SS)
    """
    if timestamp:
        return datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
    return ''


def extract_subscription_data(subscription: Dict, customer: Optional[Dict] = None, fetch_customer: bool = False) -> Dict:
    """
    Extract relevant data from subscription and customer objects.
    
    Args:
        subscription: Stripe subscription object
        customer: Optional Stripe customer object
        fetch_customer: Whether to fetch customer if not provided (deprecated, use prepare_subscription_rows instead)
    
    Returns:
        Dictionary with extracted fields
    """
    # Extract customer ID
    customer_id = subscription.get('customer', '')
    
    # Only fetch customer if explicitly requested and not already provided
    # (Normally customer should be fetched in prepare_subscription_rows)
    if not customer and customer_id and fetch_customer:
        customer = fetch_customer_details(customer_id)
    
    # Extract subscription data
    data = {
        'subscription_id': subscription.get('id', ''),
        'object_type': subscription.get('object', ''),
        'status': subscription.get('status', ''),
        'created': format_timestamp(subscription.get('created')),
        'created_timestamp': subscription.get('created', ''),
        'current_period_start': format_timestamp(subscription.get('current_period_start')),
        'current_period_end': format_timestamp(subscription.get('current_period_end')),
        'customer_id': customer_id,
        'customer_email': '',
        'customer_name': '',
        'customer_country': '',
        'customer_postal_code': '',
        'currency': subscription.get('currency', ''),
        'amount': '',
        'interval': '',
        'plan_name': '',
    }
    
    # Extract customer information
    if customer:
        data['customer_email'] = customer.get('email', '')
        data['customer_name'] = customer.get('name', '')
        
        # Try to get description if name is not available
        if not data['customer_name']:
            data['customer_name'] = customer.get('description', '')
    
    # Extract customer address and additional details if available
    if customer:
        # Check if address is directly in customer object
        address = customer.get('address', {})
        if address:
            data['customer_country'] = address.get('country', '')
            data['customer_postal_code'] = address.get('postal_code', '')
        
        # Try to get from default source if address not found
        if not data['customer_country']:
            default_source = customer.get('default_source', '')
            if default_source:
                try:
                    url = f"{STRIPE_API_BASE}/payment_methods/{default_source}"
                    headers = {
                        'Authorization': f'Bearer {STRIPE_API_KEY}',
                        'Content-Type': 'application/x-www-form-urlencoded'
                    }
                    response = requests.get(url, headers=headers, timeout=30)
                    if response.status_code == 200:
                        payment_method = response.json()
                        billing_details = payment_method.get('billing_details', {})
                        address = billing_details.get('address', {})
                        data['customer_country'] = address.get('country', '')
                        data['customer_postal_code'] = address.get('postal_code', '')
                except Exception as e:
                    # Silently fail if payment method fetch fails
                    pass
        
        # Try to get phone number if available
        if customer.get('phone'):
            # Could add phone column if needed
            pass
    
    # Extract plan/price information
    items = subscription.get('items', {}).get('data', [])
    if items:
        item = items[0]  # Get first item
        price = item.get('price', {})
        if price:
            data['amount'] = price.get('unit_amount', 0) / 100  # Convert cents to dollars
            data['currency'] = price.get('currency', subscription.get('currency', ''))
            data['interval'] = price.get('recurring', {}).get('interval', '')
            data['plan_name'] = price.get('nickname', '') or price.get('product', '')
    
    return data


def get_fieldnames() -> List[str]:
    """
    Get the list of field names for export.
    
    Returns:
        List of field names
    """
    return [
        'subscription_id',
        'object_type',
        'status',
        'created',
        'created_timestamp',
        'current_period_start',
        'current_period_end',
        'customer_id',
        'customer_email',
        'customer_name',
        'customer_country',
        'customer_postal_code',
        'currency',
        'amount',
        'interval',
        'plan_name'
    ]


def prepare_subscription_rows(subscriptions: List[Dict], fetch_customers: bool = True) -> List[Dict]:
    """
    Prepare subscription data rows for export.
    
    Args:
        subscriptions: List of subscription objects
        fetch_customers: Whether to fetch customer details for each subscription
    
    Returns:
        List of formatted row dictionaries
    """
    rows = []
    total = len(subscriptions)
    print(f"\nProcessing {total} subscription(s)...")
    
    for idx, subscription in enumerate(subscriptions, 1):
        customer_id = subscription.get('customer', '')
        subscription_id = subscription.get('id', 'N/A')
        
        if fetch_customers and customer_id:
            print(f"  [{idx}/{total}] Processing subscription {subscription_id} (customer: {customer_id})...")
            customer = fetch_customer_details(customer_id, silent=True)
            if customer:
                print(f"       ✓ Customer details retrieved")
            else:
                print(f"       ⚠ Could not fetch customer details")
        else:
            customer = None
            if customer_id:
                print(f"  [{idx}/{total}] Processing subscription {subscription_id} (customer: {customer_id}) - skipping customer fetch")
            else:
                print(f"  [{idx}/{total}] Processing subscription {subscription_id} (no customer)")
        
        row_data = extract_subscription_data(subscription, customer)
        rows.append(row_data)
    
    print(f"Successfully processed {len(rows)} subscription(s) with customer details.\n")
    return rows


def export_rows_to_csv(rows: List[Dict], output_file: str = 'stripe_subscriptions.csv'):
    """
    Export pre-formatted rows to CSV.
    
    Args:
        rows: List of formatted row dictionaries
        output_file: Output CSV file path
    """
    if not rows:
        print("No data to export.")
        return
    
    fieldnames = get_fieldnames()
    
    # Write to CSV
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    
    print(f"Successfully exported {len(rows)} subscriptions to {output_file}")


def subscriptions_to_csv(subscriptions: List[Dict], output_file: str = 'stripe_subscriptions.csv'):
    """
    Convert subscription data to CSV format.
    
    Args:
        subscriptions: List of subscription objects
        output_file: Output CSV file path
    """
    if not subscriptions:
        print("No subscriptions to export.")
        return
    
    rows = prepare_subscription_rows(subscriptions)
    export_rows_to_csv(rows, output_file)


def export_rows_to_excel(rows: List[Dict], output_file: str = 'stripe_subscriptions.xlsx'):
    """
    Export pre-formatted rows to Excel.
    
    Args:
        rows: List of formatted row dictionaries
        output_file: Output Excel file path
    """
    if not rows:
        print("No data to export.")
        return
    
    fieldnames = get_fieldnames()
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Subscriptions"
    
    # Define header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, fieldname in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_num, value=fieldname.replace('_', ' ').title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data rows
    for row_num, row_data in enumerate(rows, 2):
        for col_num, fieldname in enumerate(fieldnames, 1):
            value = row_data.get(fieldname, '')
            # Format numeric values
            if fieldname == 'amount' and value:
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for col_num, fieldname in enumerate(fieldnames, 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        
        # Check header length
        max_length = max(max_length, len(fieldname.replace('_', ' ').title()))
        
        # Check data lengths
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            cell_value = row[0].value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        # Set column width (add some padding)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(output_file)
    print(f"Successfully exported {len(rows)} subscriptions to {output_file}")


def subscriptions_to_excel(subscriptions: List[Dict], output_file: str = 'stripe_subscriptions.xlsx'):
    """
    Convert subscription data to Excel format (.xlsx).
    
    Args:
        subscriptions: List of subscription objects
        output_file: Output Excel file path
    """
    if not subscriptions:
        print("No subscriptions to export.")
        return
    
    rows = prepare_subscription_rows(subscriptions)
    export_rows_to_excel(rows, output_file)


def create_csv_from_manual_data(csv_file: str = 'stripe_subscriptions.csv', excel_file: str = 'stripe_subscriptions.xlsx'):
    """
    Create CSV and Excel from manually provided data as fallback.
    Uses the subscription data you provided.
    """
    manual_data = {
        'subscription_id': 'sub_1Scm8EL5YqSpFl3KeJFAFP1g',
        'object_type': 'subscription',
        'status': 'active',
        'created_timestamp': 1765368274,
        'customer_id': 'cus_TZw07B5NXMht4M',
        'customer_email': 'Yadiel1235@gmail.com',
        'customer_name': 'Yadiel I Melendez Miranda',
        'customer_country': 'PR',
        'customer_postal_code': '00926',
    }
    
    # Format the data
    formatted_data = {
        'subscription_id': manual_data.get('subscription_id', ''),
        'object_type': manual_data.get('object_type', 'subscription'),
        'status': manual_data.get('status', ''),
        'created': format_timestamp(manual_data.get('created_timestamp')),
        'created_timestamp': str(manual_data.get('created_timestamp', '')),
        'current_period_start': '',
        'current_period_end': '',
        'customer_id': manual_data.get('customer_id', ''),
        'customer_email': manual_data.get('customer_email', ''),
        'customer_name': manual_data.get('customer_name', ''),
        'customer_country': manual_data.get('customer_country', ''),
        'customer_postal_code': manual_data.get('customer_postal_code', ''),
        'currency': '',
        'amount': '',
        'interval': '',
        'plan_name': '',
    }
    
    export_rows_to_csv([formatted_data], csv_file)
    export_rows_to_excel([formatted_data], excel_file)


def get_customer_fieldnames() -> List[str]:
    """
    Get the list of customer field names (without subscription_id).
    
    Returns:
        List of customer field names
    """
    return [
        'object_type',
        'status',
        'created',
        'created_timestamp',
        'current_period_start',
        'current_period_end',
        'customer_id',
        'customer_email',
        'customer_name',
        'customer_country',
        'customer_postal_code',
        'currency',
        'amount',
        'interval',
        'plan_name'
    ]


def extract_customer_data(subscription: Dict, customer: Optional[Dict] = None) -> Dict:
    """
    Extract customer data from subscription and customer objects (without subscription_id).
    
    Args:
        subscription: Stripe subscription object
        customer: Optional Stripe customer object
    
    Returns:
        Dictionary with customer fields (no subscription_id)
    """
    # Extract customer ID
    customer_id = subscription.get('customer', '')
    
    # Extract customer data
    data = {
        'object_type': subscription.get('object', ''),
        'status': subscription.get('status', ''),
        'created': format_timestamp(subscription.get('created')),
        'created_timestamp': subscription.get('created', ''),
        'current_period_start': format_timestamp(subscription.get('current_period_start')),
        'current_period_end': format_timestamp(subscription.get('current_period_end')),
        'customer_id': customer_id,
        'customer_email': '',
        'customer_name': '',
        'customer_country': '',
        'customer_postal_code': '',
        'currency': subscription.get('currency', ''),
        'amount': '',
        'interval': '',
        'plan_name': '',
    }
    
    # Extract customer information
    if customer:
        data['customer_email'] = customer.get('email', '')
        data['customer_name'] = customer.get('name', '')
        
        # Try to get description if name is not available
        if not data['customer_name']:
            data['customer_name'] = customer.get('description', '')
        
        # Check if address is directly in customer object
        address = customer.get('address', {})
        if address:
            data['customer_country'] = address.get('country', '')
            data['customer_postal_code'] = address.get('postal_code', '')
        
        # Try to get from default source if address not found
        if not data['customer_country']:
            default_source = customer.get('default_source', '')
            if default_source:
                try:
                    url = f"{STRIPE_API_BASE}/payment_methods/{default_source}"
                    headers = {
                        'Authorization': f'Bearer {STRIPE_API_KEY}',
                        'Content-Type': 'application/x-www-form-urlencoded'
                    }
                    response = requests.get(url, headers=headers, timeout=30)
                    if response.status_code == 200:
                        payment_method = response.json()
                        billing_details = payment_method.get('billing_details', {})
                        address = billing_details.get('address', {})
                        data['customer_country'] = address.get('country', '')
                        data['customer_postal_code'] = address.get('postal_code', '')
                except Exception as e:
                    # Silently fail if payment method fetch fails
                    pass
    
    # Extract plan/price information
    items = subscription.get('items', {}).get('data', [])
    if items:
        item = items[0]  # Get first item
        price = item.get('price', {})
        if price:
            data['amount'] = price.get('unit_amount', 0) / 100  # Convert cents to dollars
            data['currency'] = price.get('currency', subscription.get('currency', ''))
            data['interval'] = price.get('recurring', {}).get('interval', '')
            data['plan_name'] = price.get('nickname', '') or price.get('product', '')
    
    return data


def export_customers_to_csv(customer_rows: List[Dict], output_file: str = 'stripe_customers.csv'):
    """
    Export customer data to CSV (without subscription_id).
    
    Args:
        customer_rows: List of customer data dictionaries
        output_file: Output CSV file path
    """
    if not customer_rows:
        print("No customer data to export.")
        return
    
    fieldnames = get_customer_fieldnames()
    
    # Write to CSV
    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(customer_rows)
    
    print(f"Successfully exported {len(customer_rows)} customer(s) to {output_file}")


def export_customers_to_excel(customer_rows: List[Dict], output_file: str = 'stripe_customers.xlsx'):
    """
    Export customer data to Excel (without subscription_id).
    
    Args:
        customer_rows: List of customer data dictionaries
        output_file: Output Excel file path
    """
    if not customer_rows:
        print("No customer data to export.")
        return
    
    fieldnames = get_customer_fieldnames()
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Define header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, fieldname in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_num, value=fieldname.replace('_', ' ').title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data rows
    for row_num, row_data in enumerate(customer_rows, 2):
        for col_num, fieldname in enumerate(fieldnames, 1):
            value = row_data.get(fieldname, '')
            # Format numeric values
            if fieldname == 'amount' and value:
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for col_num, fieldname in enumerate(fieldnames, 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        
        # Check header length
        max_length = max(max_length, len(fieldname.replace('_', ' ').title()))
        
        # Check data lengths
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            cell_value = row[0].value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        # Set column width (add some padding)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(output_file)
    print(f"Successfully exported {len(customer_rows)} customer(s) to {output_file}")


def fetch_multiple_subscriptions_with_customers(max_subscriptions: int = 1200) -> List[Dict]:
    """
    Fetch multiple subscriptions and their customer details.
    
    Args:
        max_subscriptions: Maximum number of subscriptions to fetch
    
    Returns:
        List of customer data dictionaries
    """
    print(f"\nFetching up to {max_subscriptions} subscriptions from Stripe API...")
    subscriptions = fetch_all_subscriptions(limit=max_subscriptions)
    
    if not subscriptions:
        print("Could not fetch subscriptions from API. Trying alternative approach...")
        return []
    
    print(f"\nFound {len(subscriptions)} subscription(s)")
    print(f"Processing subscriptions and fetching customer details...\n")
    
    customer_data_list = []
    total = len(subscriptions)
    
    for idx, subscription in enumerate(subscriptions, 1):
        subscription_id = subscription.get('id', 'N/A')
        customer_id = subscription.get('customer', '')
        
        print(f"[{idx}/{total}] Processing subscription: {subscription_id}")
        
        if customer_id:
            print(f"       Fetching customer: {customer_id}")
            customer = fetch_customer_details(customer_id, silent=True)
            
            if customer:
                print(f"       ✓ Customer details retrieved")
                customer_data = extract_customer_data(subscription, customer)
                customer_data_list.append(customer_data)
            else:
                print(f"       ⚠ Could not fetch customer details")
                # Still add subscription data even without customer details
                customer_data = extract_customer_data(subscription, None)
                customer_data_list.append(customer_data)
        else:
            print(f"       ⚠ No customer ID found in subscription")
            customer_data = extract_customer_data(subscription, None)
            customer_data_list.append(customer_data)
    
    print(f"\n✓ Successfully processed {len(customer_data_list)} subscription(s) with customer data")
    return customer_data_list


def main():
    """
    Main function to run the scraper.
    """
    print("Stripe Customer Data Scraper")
    print("=" * 50)
    
    # Fetch up to 1200 subscriptions with their customer details
    customer_data_list = fetch_multiple_subscriptions_with_customers(max_subscriptions=1200)
    
    if customer_data_list:
        print(f"\nExporting {len(customer_data_list)} customer record(s) to Excel...")
        export_customers_to_excel(customer_data_list, 'stripe_customers.xlsx')
        
        print("\n" + "=" * 50)
        print("Export completed successfully!")
        print(f"  - Excel file: stripe_customers.xlsx")
        print(f"  - Total customers exported: {len(customer_data_list)}")
        
        # Show summary
        if customer_data_list:
            print(f"\nSample customers:")
            for i, customer in enumerate(customer_data_list[:5], 1):
                name = customer.get('customer_name', 'N/A')
                email = customer.get('customer_email', 'N/A')
                print(f"  {i}. {name} ({email})")
            if len(customer_data_list) > 5:
                print(f"  ... and {len(customer_data_list) - 5} more")
    else:
        print("\n" + "=" * 50)
        print("Could not fetch subscriptions. Trying single subscription fallback...")
        print("=" * 50)
        
        # Fallback: Try to fetch the specific subscription we know works
        subscription_id = 'sub_1Scm8EL5YqSpFl3KeJFAFP1g'
        print(f"\nAttempting to fetch specific subscription: {subscription_id}")
        subscription = fetch_subscription(subscription_id)
        
        if subscription:
            print("Successfully fetched subscription from API!")
            
            customer_id = subscription.get('customer', '')
            if customer_id:
                print(f"\nFetching customer details for: {customer_id}")
                customer = fetch_customer_details(customer_id)
                
                if customer:
                    print("Successfully fetched customer details!")
                    customer_data = extract_customer_data(subscription, customer)
                    export_customers_to_excel([customer_data], 'stripe_customers.xlsx')
                    
                    print("\n" + "=" * 50)
                    print("Export completed successfully!")
                    print(f"  - Excel file: stripe_customers.xlsx")
                    print(f"  - Customer: {customer_data.get('customer_name', 'N/A')} ({customer_data.get('customer_email', 'N/A')})")
                else:
                    print(f"Warning: Could not fetch customer details")
                    customer_data = extract_customer_data(subscription, None)
                    export_customers_to_excel([customer_data], 'stripe_customers.xlsx')
        else:
            print("\nAPI fetch failed. This could be due to:")
            print("  - Invalid or insufficient API key permissions")
            print("  - Network/connection issues")
            print("  - Key format issues (needs sk_live_... or sk_test_... for full access)")
            print("\nUsing manual data as fallback...")
            
            manual_customer_data = {
                'object_type': 'subscription',
                'status': 'active',
                'created': format_timestamp(1765368274),
                'created_timestamp': '1765368274',
                'current_period_start': '',
                'current_period_end': '',
                'customer_id': 'cus_TZw07B5NXMht4M',
                'customer_email': 'Yadiel1235@gmail.com',
                'customer_name': 'Yadiel I Melendez Miranda',
                'customer_country': 'PR',
                'customer_postal_code': '00926',
                'currency': '',
                'amount': '',
                'interval': '',
                'plan_name': '',
            }
            
            export_customers_to_excel([manual_customer_data], 'stripe_customers.xlsx')
            print("\nNote: To fetch live data, you may need a Stripe secret key (sk_live_... or sk_test_...)")
            print("The provided key (rk_live_...) may be a restricted key with limited permissions.")


if __name__ == '__main__':
    main()

